from io import BytesIO
from datetime import date, datetime
from flask import render_template, request, send_file, abort
from flask_login import login_required, current_user
from app.reports import bp
from app.models import VehicleRequest, Vehicle, User, Role
from app.extensions import db
from app.decorators import role_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def _build_query(form_data):
    """Retorna query filtrada de acordo com o perfil e parâmetros."""
    query = VehicleRequest.query.join(VehicleRequest.employee).join(VehicleRequest.vehicle)

    # Coordenador só vê seus funcionários
    if current_user.role == Role.COORDINATOR:
        employee_ids = [u.id for u in current_user.employees]
        query = query.filter(VehicleRequest.employee_id.in_(employee_ids))

    # Filtros
    if form_data.get('employee_id'):
        query = query.filter(VehicleRequest.employee_id == int(form_data['employee_id']))
    if form_data.get('vehicle_id'):
        query = query.filter(VehicleRequest.vehicle_id == int(form_data['vehicle_id']))
    if form_data.get('status'):
        query = query.filter(VehicleRequest.status == form_data['status'])
    if form_data.get('date_from'):
        try:
            dt = datetime.strptime(form_data['date_from'], '%Y-%m-%d')
            query = query.filter(VehicleRequest.departure_datetime >= dt)
        except ValueError:
            pass
    if form_data.get('date_to'):
        try:
            dt = datetime.strptime(form_data['date_to'], '%Y-%m-%d')
            dt = dt.replace(hour=23, minute=59, second=59)
            query = query.filter(VehicleRequest.departure_datetime <= dt)
        except ValueError:
            pass

    return query.order_by(VehicleRequest.departure_datetime.desc())


def _employee_choices():
    if current_user.role == Role.COORDINATOR:
        return current_user.employees
    return User.query.filter_by(role=Role.EMPLOYEE, is_active=True).order_by(User.full_name).all()


@bp.route('/')
@login_required
@role_required(Role.ADMIN, Role.COORDINATOR)
def index():
    employees = _employee_choices()
    vehicles = Vehicle.query.filter_by(is_active=True).order_by(Vehicle.name).all()
    results = _build_query(request.args).all() if any(request.args.values()) else []
    return render_template(
        'reports/index.html',
        title='Relatórios',
        results=results,
        employees=employees,
        vehicles=vehicles,
        args=request.args,
    )


@bp.route('/exportar/excel')
@login_required
@role_required(Role.ADMIN, Role.COORDINATOR)
def export_excel():
    results = _build_query(request.args).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Saídas de Veículos'

    header_fill = PatternFill(start_color='212529', end_color='212529', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        '#', 'Funcionário', 'Veículo', 'Placa',
        'Saída Prevista', 'Retorno Previsto',
        'Saída Real', 'Retorno Real',
        'Motivo', 'Retorna?', 'Status', 'Observação',
    ]
    col_widths = [5, 28, 20, 12, 18, 18, 18, 18, 35, 10, 12, 30]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20

    status_map = {'pending': 'Pendente', 'approved': 'Aprovado', 'rejected': 'Recusado'}

    for row_idx, req in enumerate(results, start=2):
        row_data = [
            req.id,
            req.employee.full_name,
            req.vehicle.name,
            req.vehicle.plate,
            req.departure_datetime.strftime('%d/%m/%Y %H:%M'),
            req.expected_return_datetime.strftime('%d/%m/%Y %H:%M'),
            req.actual_departure_datetime.strftime('%d/%m/%Y %H:%M') if req.actual_departure_datetime else '—',
            req.actual_return_datetime.strftime('%d/%m/%Y %H:%M') if req.actual_return_datetime else '—',
            req.reason,
            'Sim' if req.returns_to_company else 'Não',
            status_map.get(req.status, req.status),
            req.coordinator_notes or '',
        ]
        fill_color = 'F8F9FA' if row_idx % 2 == 0 else 'FFFFFF'
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'relatorio_saidas_{date.today().isoformat()}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@bp.route('/exportar/pdf')
@login_required
@role_required(Role.ADMIN, Role.COORDINATOR)
def export_pdf():
    results = _build_query(request.args).all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=14, spaceAfter=6)
    subtitle_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, textColor=colors.grey)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10)

    status_map = {'pending': 'Pendente', 'approved': 'Aprovado', 'rejected': 'Recusado'}

    header_row = [
        Paragraph('<b>#</b>', cell_style),
        Paragraph('<b>Funcionário</b>', cell_style),
        Paragraph('<b>Veículo</b>', cell_style),
        Paragraph('<b>Saída Prevista</b>', cell_style),
        Paragraph('<b>Retorno Previsto</b>', cell_style),
        Paragraph('<b>Saída Real</b>', cell_style),
        Paragraph('<b>Retorno Real</b>', cell_style),
        Paragraph('<b>Status</b>', cell_style),
        Paragraph('<b>Motivo</b>', cell_style),
    ]

    data = [header_row]
    for req in results:
        data.append([
            Paragraph(str(req.id), cell_style),
            Paragraph(req.employee.full_name, cell_style),
            Paragraph(f'{req.vehicle.name}\n{req.vehicle.plate}', cell_style),
            Paragraph(req.departure_datetime.strftime('%d/%m/%Y\n%H:%M'), cell_style),
            Paragraph(req.expected_return_datetime.strftime('%d/%m/%Y\n%H:%M'), cell_style),
            Paragraph(req.actual_departure_datetime.strftime('%d/%m/%Y\n%H:%M')
                      if req.actual_departure_datetime else '—', cell_style),
            Paragraph(req.actual_return_datetime.strftime('%d/%m/%Y\n%H:%M')
                      if req.actual_return_datetime else '—', cell_style),
            Paragraph(status_map.get(req.status, req.status), cell_style),
            Paragraph(req.reason[:80], cell_style),
        ])

    col_widths_pdf = [1*cm, 4.5*cm, 3.5*cm, 3*cm, 3*cm, 3*cm, 3*cm, 2.5*cm, 5*cm]

    table = Table(data, colWidths=col_widths_pdf, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#212529')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
    ]))

    elements = [
        Paragraph('Relatório de Saídas de Veículos', title_style),
        Paragraph(f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} — {len(results)} registro(s)',
                  subtitle_style),
        Spacer(1, 0.4 * cm),
        table,
    ]

    doc.build(elements)
    buffer.seek(0)
    filename = f'relatorio_saidas_{date.today().isoformat()}.pdf'
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
